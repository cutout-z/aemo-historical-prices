"""Excel output generation — per-region workbooks with summary, data, and heatmap."""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from . import config

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CARBON_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
MONEY_FORMAT = '#,##0.00'


def generate_all_workbooks(summary: pd.DataFrame, output_dir: str):
    """Generate one .xlsx workbook per region from the summary DataFrame."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    for region in config.REGIONS:
        friendly = config.REGION_NAMES[region]
        region_data = summary[summary["region"] == region].copy()
        if region_data.empty:
            logger.warning(f"No data for {region}, skipping workbook")
            continue

        region_data = region_data.sort_values("year_month").reset_index(drop=True)
        filepath = output_path / f"{friendly}_historical_prices.xlsx"

        _write_region_workbook(region_data, friendly, filepath)
        logger.info(f"Written {filepath.name}")


def _format_month(year_month: str) -> str:
    """Convert '2003-07' to 'Jul 2003'."""
    dt = pd.Timestamp(year_month + "-01")
    return dt.strftime("%b %Y")


def _write_region_workbook(data: pd.DataFrame, region_name: str, filepath: Path):
    """Write a 3-sheet workbook for a single region."""
    wb = Workbook()

    _write_summary_sheet(wb, data, region_name)
    _write_data_sheet(wb, data, region_name)
    _write_heatmap_sheet(wb, data, region_name)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)


def _write_summary_sheet(wb: Workbook, data: pd.DataFrame, region_name: str):
    """Sheet 1: Rolling averages for each period."""
    ws = wb.create_sheet(title="Summary")

    # Title
    ws.cell(row=1, column=1, value=f"{region_name} — Historical Price Summary").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Rolling averages ($/MWh)").font = Font(size=11, italic=True)

    # Headers
    headers = ["Period", "RRP (Nominal)", "Peak RRP (Nominal)", "RRP (Real)", "Peak RRP (Real)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Calculate rolling averages
    row_idx = 5
    for period in config.ROLLING_PERIODS:
        months_needed = period * 12
        if len(data) < months_needed:
            continue

        recent = data.tail(months_needed)
        ws.cell(row=row_idx, column=1, value=f"{period}-Year Average").border = THIN_BORDER

        for col_idx, col in enumerate(["rrp_nominal", "peak_rrp_nominal", "rrp_real", "peak_rrp_real"], 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=round(recent[col].mean(), 2))
            cell.number_format = MONEY_FORMAT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        row_idx += 1

    # Latest data info
    latest = data.iloc[-1]
    ws.cell(row=row_idx + 1, column=1, value=f"Data through: {_format_month(latest['year_month'])}").font = Font(
        size=10, italic=True
    )

    # Column widths
    ws.column_dimensions["A"].width = 20
    for col_idx in range(2, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def _write_data_sheet(wb: Workbook, data: pd.DataFrame, region_name: str):
    """Sheet 2: Full monthly time series."""
    ws = wb.create_sheet(title="Monthly Data")

    headers = [
        "Month", "RRP (Nominal)", "Peak RRP (Nominal)",
        "RRP (Real)", "Peak RRP (Real)", "Carbon Tax"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        month_cell = ws.cell(row=row_idx, column=1, value=_format_month(row["year_month"]))
        month_cell.border = THIN_BORDER

        for col_idx, col in enumerate(["rrp_nominal", "peak_rrp_nominal", "rrp_real", "peak_rrp_real"], 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col])
            cell.number_format = MONEY_FORMAT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        carbon_cell = ws.cell(row=row_idx, column=6, value="Yes" if row.get("carbon_flag") else "")
        carbon_cell.alignment = Alignment(horizontal="center")
        carbon_cell.border = THIN_BORDER

        if row.get("carbon_flag"):
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = CARBON_FILL

    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.freeze_panes = "A2"


def _write_heatmap_sheet(wb: Workbook, data: pd.DataFrame, region_name: str):
    """Sheet 3: Monthly data with conditional formatting heatmap."""
    ws = wb.create_sheet(title="Heatmap")

    headers = ["Month", "RRP (Nominal)", "Peak RRP (Nominal)", "RRP (Real)", "Peak RRP (Real)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    num_rows = len(data)
    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=_format_month(row["year_month"])).border = THIN_BORDER
        for col_idx, col in enumerate(["rrp_nominal", "peak_rrp_nominal", "rrp_real", "peak_rrp_real"], 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col])
            cell.number_format = MONEY_FORMAT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Apply colour scale per column (green=low → yellow=mid → red=high)
    if num_rows > 0:
        for col_idx in range(2, 6):
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{num_rows + 1}"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="min", start_color="63BE7B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="F8696B",
                ),
            )

    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.freeze_panes = "A2"
